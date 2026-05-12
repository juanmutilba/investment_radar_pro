from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

fill_green_dark = PatternFill(fill_type='solid', start_color='00B050', end_color='00B050')
fill_green_light = PatternFill(fill_type='solid', start_color='C6EFCE', end_color='C6EFCE')
fill_yellow = PatternFill(fill_type='solid', start_color='FFF2CC', end_color='FFF2CC')
fill_orange = PatternFill(fill_type='solid', start_color='F4B183', end_color='F4B183')
fill_red = PatternFill(fill_type='solid', start_color='FFC7CE', end_color='FFC7CE')
fill_gray = PatternFill(fill_type='solid', start_color='D9D9D9', end_color='D9D9D9')


HEADER_COMMENTS = {
    'TechScore': 'Indicadores técnicos\n\nRSI\nMA50\nMA200\nMACD\nPullback\n\nMáximo 7 puntos',
    'FundScore': 'Resumen fundamental\n\nPE\nEBITDA\nNet Income\nDebt/Equity\nUpside\n\nMáximo 5 puntos',
    'RiskScore': 'Beta\n\n+1: beta ideal\n0: beta neutral\n-1: beta extrema\n\nMáximo 1\nMínimo -1',
    'TotalScore': 'Score total\n\nTechScore + FundScore + RiskScore\n\nMáximo actual: 13',
    'SignalState': 'Clasificación operativa automática\nbasada en score y condiciones técnicas',
    'PrioridadRadar': 'Prioridad diaria de seguimiento\nbasada en score y evolución',
}


def add_header_comments(ws):
    for cell in ws[1]:
        if cell.value in HEADER_COMMENTS:
            cell.comment = Comment(HEADER_COMMENTS[cell.value], 'Radar')


def apply_signal_fill(cell):
    value = str(cell.value).strip() if cell.value is not None else ''
    if value == 'COMPRA PRIORITARIA':
        cell.fill = fill_green_dark
    elif value == 'COMPRA POTENCIAL':
        cell.fill = fill_green_light
    elif value == 'SEGUIMIENTO':
        cell.fill = fill_yellow
    elif value == 'TOMA DE GANANCIA':
        cell.fill = fill_orange
    elif value in ['SOBREEXTENDIDA', 'DEBILITÁNDOSE']:
        cell.fill = fill_red
    elif value == 'EVITAR':
        cell.fill = fill_gray


def apply_evolution_fill(cell):
    value = str(cell.value).strip() if cell.value is not None else ''
    if value == 'MEJORANDO':
        cell.fill = fill_green_light
    elif value == 'DETERIORANDO':
        cell.fill = fill_red
    elif value == 'CAMBIO DE ESTADO':
        cell.fill = fill_yellow
    elif value == 'NUEVA INCORPORACIÓN':
        cell.fill = fill_green_dark
    elif value in ['SIN CAMBIOS', 'SIN HISTORIAL']:
        cell.fill = fill_gray


def apply_score_fill(cell):
    try:
        value = float(cell.value)
    except Exception:
        return
    if value >= 10:
        cell.fill = fill_green_dark
    elif value >= 8:
        cell.fill = fill_green_light
    elif value >= 5:
        cell.fill = fill_yellow
    else:
        cell.fill = fill_red


def apply_change_score_fill(cell):
    try:
        value = float(cell.value)
    except Exception:
        return
    if value >= 2:
        cell.fill = fill_green_dark
    elif value > 0:
        cell.fill = fill_green_light
    elif value == 0:
        cell.fill = fill_gray
    elif value <= -2:
        cell.fill = fill_red
    else:
        cell.fill = fill_orange


def apply_priority_fill(cell):
    value = str(cell.value).strip() if cell.value is not None else ''
    if value == 'ALTA':
        cell.fill = fill_green_dark
    elif value == 'MEDIA':
        cell.fill = fill_green_light
    elif value == 'BAJA':
        cell.fill = fill_yellow
    elif value == 'IGNORAR':
        cell.fill = fill_gray


def apply_risk_fill(cell):
    value = str(cell.value).strip() if cell.value is not None else ''
    if value == 'BALANCEADO':
        cell.fill = fill_green_light
    elif value == 'AGRESIVO':
        cell.fill = fill_yellow
    elif value == 'DEFENSIVO':
        cell.fill = fill_gray
    elif value == 'ESPECULATIVO':
        cell.fill = fill_red


def format_workbook(wb):
    # Evita recorrer decenas de miles de filas por columna (colgaba el scan en hojas Universo).
    _MAX_WIDTH_SCAN_ROWS = 800
    _MAX_COL_WIDTH = 55
    _MIN_COL_WIDTH = 10
    _MAX_ROWS_CONDITIONAL_FORMAT = 12000

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        add_header_comments(ws)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        max_scan = min(ws.max_row, _MAX_WIDTH_SCAN_ROWS)
        if ws.max_column and ws.max_row:
            for col_cells in ws.iter_cols(
                min_row=1, max_row=max_scan, min_col=1, max_col=ws.max_column
            ):
                col_letter = col_cells[0].column_letter
                max_length = 0
                for cell in col_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                width = min(max_length + 2, _MAX_COL_WIDTH)
                width = max(width, _MIN_COL_WIDTH)
                ws.column_dimensions[col_letter].width = width

        header_map = {cell.value: cell.column for cell in ws[1]}

        last_fmt_row = min(ws.max_row, _MAX_ROWS_CONDITIONAL_FORMAT)
        for row in range(2, last_fmt_row + 1):
            if 'SignalState' in header_map:
                apply_signal_fill(ws.cell(row=row, column=header_map['SignalState']))
            if 'Evolucion' in header_map:
                apply_evolution_fill(ws.cell(row=row, column=header_map['Evolucion']))
            if 'TotalScore' in header_map:
                apply_score_fill(ws.cell(row=row, column=header_map['TotalScore']))
            if 'CambioScore' in header_map:
                apply_change_score_fill(ws.cell(row=row, column=header_map['CambioScore']))
            if 'PrioridadRadar' in header_map:
                apply_priority_fill(ws.cell(row=row, column=header_map['PrioridadRadar']))
            if 'RiskProfile' in header_map:
                apply_risk_fill(ws.cell(row=row, column=header_map['RiskProfile']))
